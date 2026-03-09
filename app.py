import logging
import warnings
from flask import Flask, render_template, request, jsonify, session, redirect, url_for
from flask_cors import CORS
import firebase_admin
from firebase_admin import credentials, firestore
import anthropic
import openpyxl
import os
import json
import requests
import threading
import time
from datetime import datetime
from dotenv import load_dotenv
import io

load_dotenv()

# Suprimir warnings y logs innecesarios
warnings.filterwarnings("ignore")
logging.getLogger("werkzeug").setLevel(logging.ERROR)
logging.getLogger("google").setLevel(logging.ERROR)
logging.getLogger("urllib3").setLevel(logging.ERROR)

app = Flask(__name__)
app.secret_key = os.getenv('SECRET_KEY', 'clave_secreta_123')
CORS(app)

# Firebase init
cred = credentials.Certificate(os.getenv('FIREBASE_CREDENTIALS', 'firebase_credentials.json'))
firebase_admin.initialize_app(cred)
db = firestore.client()

# Claude init
claude = anthropic.Anthropic(api_key=os.getenv('ANTHROPIC_API_KEY'))

UPLOAD_FOLDER = 'uploads'
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# URLs de Drive y Forms
EXCEL_DOWNLOAD_URL = "https://docs.google.com/spreadsheets/d/1JAjlUyYLGd636Dayzt0X7DitegV61xeI/export?format=xlsx"
SEB_INSTALLER_URL = "https://drive.google.com/uc?export=download&id=1JI1qX2L2uoKs2J1wTl4DYo0IWFCGDKqR"
SEB_CONFIG_URL = "https://drive.google.com/uc?export=download&id=18pivzPn-JTKXmtpSu7ncdzxOOSTlBvBj"
SHEETS_CSV_URL = "https://docs.google.com/spreadsheets/d/1OKrBXOXssQv8gMoBYLqBfF0QENA4NRp9qL1_2J2GIlY/export?format=csv&gid=207460306"

# Usuarios hardcodeados
DOCENTE = {"email": "docente@evaluacion.com", "password": "profe2024", "rol": "docente"}


# ==================== CACHE DE ALUMNOS ====================

_cache_alumnos = {}
_cache_timestamp = 0
_cache_lock = threading.Lock()

def get_alumnos_cache():
    """Devuelve alumnos desde memoria. Refresca cada 5 minutos desde Firebase."""
    global _cache_alumnos, _cache_timestamp
    ahora = time.time()
    with _cache_lock:
        if ahora - _cache_timestamp > 300:  # refrescar cada 5 minutos
            print("[Cache] Actualizando cache de alumnos desde Firebase...")
            try:
                docs = db.collection('alumnos').get()
                _cache_alumnos = {}
                for doc in docs:
                    d = doc.to_dict()
                    d['_doc_id'] = doc.id
                    _cache_alumnos[d['codigo']] = d
                _cache_timestamp = ahora
                print(f"[Cache] {len(_cache_alumnos)} alumnos cargados en memoria")
            except Exception as e:
                print(f"[Cache] Error al cargar alumnos: {e}")
    return _cache_alumnos

def invalidar_cache():
    """Fuerza recarga del cache en la próxima consulta."""
    global _cache_timestamp
    with _cache_lock:
        _cache_timestamp = 0

def precargar_cache():
    """Precarga el cache al iniciar el servidor."""
    time.sleep(3)  # esperar que Firebase conecte
    get_alumnos_cache()
    print("[Cache] Cache precargado al inicio")


# ==================== CONFIGURACIÓN DEL EXAMEN ====================

_cache_config = {}
_cache_config_timestamp = 0

def get_config_examen():
    global _cache_config, _cache_config_timestamp
    ahora = time.time()
    if ahora - _cache_config_timestamp > 30:  # refrescar cada 30 segundos
        try:
            doc = db.collection("config").document("examen").get()
            if doc.exists:
                _cache_config = doc.to_dict()
                _cache_config_timestamp = ahora
        except:
            pass
    if _cache_config:
        return _cache_config
    return {"activo": False, "hora_inicio": None, "hora_fin": None, "titulo": "Examen Final Excel 365", "calificacion_iniciada": False, "calificacion_completada": False}

def set_config_examen(data):
    global _cache_config_timestamp
    db.collection("config").document("examen").set(data, merge=True)
    _cache_config_timestamp = 0  # invalidar cache de config

def examen_activo():
    config = get_config_examen()
    if not config.get("activo"):
        return False
    ahora = datetime.now().isoformat()
    inicio = config.get("hora_inicio", "")
    fin = config.get("hora_fin", "")
    if inicio and fin:
        return inicio <= ahora <= fin
    return config.get("activo", False)

def examen_cerrado():
    config = get_config_examen()
    fin = config.get("hora_fin", "")
    if not fin:
        return False
    return datetime.now().isoformat() > fin


# ==================== RUTAS PRINCIPALES ====================

@app.route('/')
def index():
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        data = request.json
        codigo = data.get('codigo', '').strip()
        password = data.get('password', '').strip()

        # Verificar si es docente
        if codigo == DOCENTE['email'] and password == DOCENTE['password']:
            session['usuario'] = codigo
            session['rol'] = 'docente'
            return jsonify({'success': True, 'rol': 'docente'})

        # Verificar alumno desde CACHE (sin consultar Firebase)
        alumnos = get_alumnos_cache()
        alumno = alumnos.get(codigo)
        if alumno and alumno.get('password') == password:
            session['usuario'] = codigo
            session['rol'] = 'alumno'
            session['nombre'] = alumno.get('nombre', '')
            return jsonify({'success': True, 'rol': 'alumno'})

        return jsonify({'success': False, 'mensaje': 'Credenciales incorrectas'})

    return render_template('login.html')

@app.route('/panel-alumno')
def panel_alumno():
    if session.get('rol') != 'alumno':
        return redirect(url_for('login'))
    return render_template('panel_student.html',
                           nombre=session.get('nombre', ''),
                           codigo=session.get('usuario', ''),
                           excel_url=EXCEL_DOWNLOAD_URL,
                           seb_installer_url=SEB_INSTALLER_URL,
                           seb_config_url=SEB_CONFIG_URL)

@app.route('/panel-docente')
def panel_docente():
    if session.get('rol') != 'docente':
        return redirect(url_for('login'))
    return render_template('panel_instructor.html')

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

@app.route("/examen-entregado")
def examen_entregado():
    return render_template("examen_entregado.html")


# ==================== API CONFIG EXAMEN ====================

@app.route('/api/config-examen', methods=['GET'])
def get_config():
    if session.get('rol') != 'docente':
        return jsonify({'error': 'No autorizado'}), 401
    return jsonify(get_config_examen())

@app.route('/api/config-examen', methods=['POST'])
def set_config():
    if session.get('rol') != 'docente':
        return jsonify({'error': 'No autorizado'}), 401
    data = request.json
    set_config_examen({
        'activo': data.get('activo', False),
        'hora_inicio': data.get('hora_inicio'),
        'hora_fin': data.get('hora_fin'),
        'titulo': data.get('titulo', 'Examen Final Excel 365'),
        'calificacion_iniciada': False,
        'calificacion_completada': False
    })
    print(f"[Config] Examen configurado -> Inicio: {data.get('hora_inicio')} | Fin: {data.get('hora_fin')}")
    return jsonify({'success': True})

@app.route('/api/estado-examen', methods=['GET'])
def estado_examen():
    config = get_config_examen()
    return jsonify({
        'activo': examen_activo(),
        'cerrado': examen_cerrado(),
        'hora_fin': config.get('hora_fin'),
        'titulo': config.get('titulo', 'Examen')
    })

@app.route('/api/iniciar-calificacion', methods=['POST'])
def iniciar_calificacion():
    if session.get('rol') != 'docente':
        return jsonify({'error': 'No autorizado'}), 401
    set_config_examen({'calificacion_iniciada': True})
    threading.Thread(target=calificar_todos_batch).start()
    return jsonify({'success': True})

def calificar_todos_batch():
    print("[Batch] Iniciando calificacion masiva...")
    alumnos = db.collection('alumnos').get()
    total = 0
    for doc in alumnos:
        d = doc.to_dict()
        if d.get('entrego') and d.get('nota_final') is None:
            filepath = os.path.join(UPLOAD_FOLDER, d.get('archivo', ''))
            if os.path.exists(filepath):
                resultado = calificar_con_ia(filepath, d['codigo'], d['nombre'])
                db.collection('alumnos').document(doc.id).update({
                    'nota_final': resultado.get('nota_final'),
                    'notas_detalle': {k: v for k, v in resultado.items() if k.startswith('ejercicio')},
                    'retroalimentacion': resultado.get('retroalimentacion', '')
                })
                total += 1
                print(f"[Batch] OK {d['nombre']} -> {resultado.get('nota_final')}")
    set_config_examen({'calificacion_completada': True})
    invalidar_cache()
    print(f"[Batch] Completado: {total} alumnos calificados")

def monitor_cierre_examen():
    while True:
        try:
            config = get_config_examen()
            if examen_cerrado() and config.get('activo') and not config.get('calificacion_iniciada'):
                print("[Monitor] Examen cerrado. Iniciando calificacion automatica...")
                set_config_examen({'calificacion_iniciada': True, 'activo': False})
                threading.Thread(target=calificar_todos_batch).start()
        except:
            pass
        time.sleep(30)


# ==================== API ALUMNOS ====================

@app.route('/api/alumnos', methods=['GET'])
def get_alumnos():
    if session.get('rol') != 'docente':
        return jsonify({'error': 'No autorizado'}), 401
    alumnos = []
    for doc in db.collection('alumnos').get():
        alumnos.append(doc.to_dict())
    return jsonify(alumnos)

@app.route('/api/alumnos/registrar', methods=['POST'])
def registrar_alumno():
    data = request.json
    db.collection('alumnos').add({
        'codigo': data['codigo'],
        'nombre': data['nombre'],
        'password': data['password'],
        'entrego': False,
        'nota_final': None,
        'fecha_entrega': None
    })
    invalidar_cache()  # refrescar cache al registrar nuevo alumno
    return jsonify({'success': True})


# ==================== API EXAMEN ====================

@app.route('/api/subir-examen', methods=['POST'])
def subir_examen():
    if 'archivo' not in request.files:
        return jsonify({'error': 'No se envió archivo'}), 400

    archivo = request.files['archivo']
    codigo = request.form.get('codigo', session.get('usuario', 'desconocido'))
    nombre = request.form.get('nombre', session.get('nombre', ''))

    filename = f"{codigo}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = os.path.join(UPLOAD_FOLDER, filename)
    archivo.save(filepath)

    resultado = calificar_con_ia(filepath, codigo, nombre)

    alumnos_ref = db.collection('alumnos').where('codigo', '==', codigo).get()
    if alumnos_ref:
        doc_id = alumnos_ref[0].id
        db.collection('alumnos').document(doc_id).update({
            'entrego': True,
            'nota_final': resultado['nota_final'],
            'notas_detalle': resultado.get('detalle', {}),
            'retroalimentacion': resultado['retroalimentacion'],
            'fecha_entrega': datetime.now().isoformat(),
            'archivo': filename
        })
    invalidar_cache()
    return jsonify({'success': True, 'resultado': resultado})


# ==================== CALIFICACIÓN CON IA ====================

def calificar_con_ia(filepath, codigo, nombre):
    MODO_DEMO = True

    if MODO_DEMO:
        return calificar_demo(nombre)

    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        contenido = extraer_contenido_excel(wb)

        prompt = f"""Eres un docente experto en Microsoft Excel 365 y Macros VBA.
        
Debes calificar el siguiente examen de Excel del alumno {nombre} (código: {codigo}).

El examen tiene 5 ejercicios, cada uno vale 20 puntos (total 100 puntos):
- Ejercicio 1: Macros y VBA (crear macro, colorear encabezados, gráfico con botón)
- Ejercicio 2: Fórmulas avanzadas
- Ejercicio 3: Tablas dinámicas
- Ejercicio 4: Formato condicional
- Ejercicio 5: Funciones de búsqueda

Contenido del examen entregado:
{contenido}

Responde SOLO en formato JSON así:
{{
    "ejercicio_1": <nota 0-20>,
    "ejercicio_2": <nota 0-20>,
    "ejercicio_3": <nota 0-20>,
    "ejercicio_4": <nota 0-20>,
    "ejercicio_5": <nota 0-20>,
    "nota_final": <promedio>,
    "retroalimentacion": "<comentario general del desempeño>"
}}"""

        response = claude.messages.create(
            model="claude-sonnet-4-20250514",
            max_tokens=1000,
            messages=[{"role": "user", "content": prompt}]
        )

        texto = response.content[0].text.strip()
        texto = texto.replace('```json', '').replace('```', '').strip()
        resultado = json.loads(texto)
        return resultado

    except Exception as e:
        return {
            "ejercicio_1": 0, "ejercicio_2": 0, "ejercicio_3": 0,
            "ejercicio_4": 0, "ejercicio_5": 0,
            "nota_final": 0,
            "retroalimentacion": f"Error al calificar: {str(e)}"
        }


def calificar_demo(nombre):
    import random
    e1 = random.randint(12, 20)
    e2 = random.randint(10, 20)
    e3 = random.randint(11, 20)
    e4 = random.randint(13, 20)
    e5 = random.randint(10, 20)
    nota_final = round((e1 + e2 + e3 + e4 + e5) / 5, 1)
    comentarios = [
        "Buen desempeño general. Se recomienda reforzar fórmulas avanzadas.",
        "Excelente manejo de macros VBA. Mejorar tablas dinámicas.",
        "Dominio aceptable del contenido. Practicar más funciones de búsqueda.",
        "Muy buen trabajo en formato condicional. Revisar ejercicio de macros.",
    ]
    return {
        "ejercicio_1": e1, "ejercicio_2": e2, "ejercicio_3": e3,
        "ejercicio_4": e4, "ejercicio_5": e5,
        "nota_final": nota_final,
        "retroalimentacion": f"[DEMO] {random.choice(comentarios)}"
    }

def extraer_contenido_excel(wb):
    contenido = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        contenido.append(f"\n=== Hoja: {sheet_name} ===")
        for row in ws.iter_rows(min_row=1, max_row=50, values_only=True):
            fila = [str(c) if c is not None else '' for c in row]
            if any(fila):
                contenido.append(' | '.join(fila))
    return '\n'.join(contenido)


# ==================== API RESULTADOS ====================

@app.route('/api/resetear-alumno', methods=['POST'])
def resetear_alumno():
    if session.get('rol') != 'docente':
        return jsonify({'error': 'No autorizado'}), 401
    codigo = request.json.get('codigo')
    docs = db.collection('alumnos').where('codigo', '==', codigo).get()
    if docs:
        db.collection('alumnos').document(docs[0].id).update({
            'entrego': False,
            'nota_final': None,
            'notas_detalle': {},
            'retroalimentacion': ''
        })
    invalidar_cache()
    return jsonify({'success': True})

@app.route('/api/resultados', methods=['GET'])
def get_resultados():
    if session.get('rol') != 'docente':
        return jsonify({'error': 'No autorizado'}), 401
    resultados = []
    for doc in db.collection('alumnos').get():
        d = doc.to_dict()
        resultados.append({
            'codigo': d.get('codigo', ''),
            'nombre': d.get('nombre', ''),
            'entrego': d.get('entrego', False),
            'nota_final': d.get('nota_final'),
            'notas_detalle': d.get('notas_detalle', {}),
            'retroalimentacion': d.get('retroalimentacion', ''),
            'fecha_entrega': d.get('fecha_entrega', '')
        })
    return jsonify(resultados)

@app.route('/api/mi-nota', methods=['GET'])
def mi_nota():
    codigo = request.args.get('codigo', session.get('usuario', ''))
    if not codigo:
        return jsonify({'error': 'No autorizado'}), 401
    # Buscar en cache primero
    alumnos = get_alumnos_cache()
    alumno = alumnos.get(codigo)
    if not alumno:
        return jsonify({'entrego': False, 'nota_final': None})
    return jsonify({
        'entrego': alumno.get('entrego', False),
        'nota_final': alumno.get('nota_final'),
        'notas_detalle': alumno.get('notas_detalle', {}),
        'retroalimentacion': alumno.get('retroalimentacion', ''),
        'fecha_entrega': alumno.get('fecha_entrega', '')
    })

@app.route('/api/estadisticas', methods=['GET'])
def get_estadisticas():
    if session.get('rol') != 'docente':
        return jsonify({'error': 'No autorizado'}), 401
    alumnos = db.collection('alumnos').get()
    total = 0
    entregaron = 0
    notas = []
    for doc in alumnos:
        d = doc.to_dict()
        total += 1
        if d.get('entrego'):
            entregaron += 1
            if d.get('nota_final') is not None:
                notas.append(d['nota_final'])
    promedio = sum(notas) / len(notas) if notas else 0
    return jsonify({
        'total': total,
        'entregaron': entregaron,
        'pendientes': total - entregaron,
        'promedio_general': round(promedio, 2),
        'nota_maxima': max(notas) if notas else 0,
        'nota_minima': min(notas) if notas else 0
    })


# ==================== PROCESAMIENTO AUTOMÁTICO DE FORMS ====================

def procesar_respuestas_forms():
    import csv
    try:
        res = requests.get(SHEETS_CSV_URL, timeout=30)
        if res.status_code != 200:
            print(f"[Forms] Error al leer Sheets: {res.status_code}")
            return

        lineas = res.content.decode('utf-8').splitlines()
        reader = csv.reader(lineas)
        filas = list(reader)

        if len(filas) < 2:
            print("[Forms] Sin respuestas nuevas")
            return

        for fila in filas[1:]:
            if len(fila) < 4:
                continue
            timestamp = fila[0].strip()
            nombre    = fila[1].strip()
            dni       = fila[2].strip()
            file_url  = fila[3].strip()

            if not dni or not file_url:
                continue

            ya_procesado = db.collection('alumnos').where('codigo', '==', dni).where('entrego', '==', True).get()
            if ya_procesado:
                continue

            print(f"[Forms] Procesando: {nombre} ({dni})")
            file_id = extraer_id_drive(file_url)
            if not file_id:
                print(f"[Forms] No se pudo extraer ID del archivo: {file_url}")
                continue

            download_url = f"https://drive.google.com/uc?export=download&id={file_id}"
            file_res = requests.get(download_url, timeout=60)
            if file_res.status_code != 200:
                print(f"[Forms] Error al descargar archivo de {dni}")
                continue

            filename = f"{dni}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            filepath = os.path.join(UPLOAD_FOLDER, filename)
            with open(filepath, 'wb') as f:
                f.write(file_res.content)

            resultado = calificar_con_ia(filepath, dni, nombre)
            print(f"[Forms] Nota de {nombre}: {resultado.get('nota_final')}")

            alumnos_ref = db.collection('alumnos').where('codigo', '==', dni).get()
            if alumnos_ref:
                doc_id = alumnos_ref[0].id
                db.collection('alumnos').document(doc_id).update({
                    'entrego': True,
                    'nota_final': resultado.get('nota_final'),
                    'notas_detalle': {k: v for k, v in resultado.items() if k.startswith('ejercicio')},
                    'retroalimentacion': resultado.get('retroalimentacion', ''),
                    'fecha_entrega': timestamp,
                    'archivo': filename
                })
            else:
                db.collection('alumnos').add({
                    'codigo': dni,
                    'nombre': nombre,
                    'password': dni,
                    'entrego': True,
                    'nota_final': resultado.get('nota_final'),
                    'notas_detalle': {k: v for k, v in resultado.items() if k.startswith('ejercicio')},
                    'retroalimentacion': resultado.get('retroalimentacion', ''),
                    'fecha_entrega': timestamp,
                    'archivo': filename
                })
            invalidar_cache()
            print(f"[Forms] ✅ {nombre} procesado correctamente")

    except Exception as e:
        print(f"[Forms] Error general: {e}")


def extraer_id_drive(url):
    import re
    patrones = [
        r'/file/d/([a-zA-Z0-9_-]+)',
        r'id=([a-zA-Z0-9_-]+)',
        r'/open\?id=([a-zA-Z0-9_-]+)'
    ]
    for patron in patrones:
        match = re.search(patron, url)
        if match:
            return match.group(1)
    return None

def monitor_forms():
    print("[Monitor] Iniciando monitoreo automático de Forms...")
    while True:
        procesar_respuestas_forms()
        time.sleep(120)

@app.route('/api/procesar-forms', methods=['POST'])
def procesar_forms_manual():
    if session.get('rol') != 'docente':
        return jsonify({'error': 'No autorizado'}), 401
    threading.Thread(target=procesar_respuestas_forms).start()
    return jsonify({'success': True, 'mensaje': 'Procesando respuestas del Forms...'})


if __name__ == '__main__':
    monitor = threading.Thread(target=monitor_forms, daemon=True)
    monitor2 = threading.Thread(target=monitor_cierre_examen, daemon=True)
    monitor3 = threading.Thread(target=precargar_cache, daemon=True)
    monitor2.start()
    monitor.start()
    monitor3.start()
    app.run(debug=True, port=5000)

# ==================== INICIO CON GUNICORN ====================
# Precargar cache cuando inicia gunicorn
_precarga = threading.Thread(target=precargar_cache, daemon=True)
_precarga.start()
_monitor_cierre = threading.Thread(target=monitor_cierre_examen, daemon=True)
_monitor_cierre.start()
_monitor_forms = threading.Thread(target=monitor_forms, daemon=True)
_monitor_forms.start()